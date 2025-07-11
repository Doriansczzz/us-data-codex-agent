import pandas as pd
import requests
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl.drawing.image

def fetch_population(state_code="06"):
    url = f"https://api.census.gov/data/2020/dec/pl?get=NAME,P1_001N&for=place:*&in=state:{state_code}"
    r = requests.get(url)
    data = r.json()
    df = pd.DataFrame(data[1:], columns=data[0])
    df.rename(columns={"NAME": "City", "P1_001N": "Population"}, inplace=True)
    df["Population"] = pd.to_numeric(df["Population"])
    return df

def fetch_income():
    url = "https://api.census.gov/data/2021/acs/acs5/subject?get=NAME,S1901_C01_012E&for=place:*"
    r = requests.get(url)
    data = r.json()
    df = pd.DataFrame(data[1:], columns=["City", "Median_Income", "State_Code", "Place_Code"])
    df["Median_Income"] = pd.to_numeric(df["Median_Income"], errors="coerce")
    return df

def fetch_crime(state_abbr="CA"):
    # Placeholder dataset from FBI UCR – ideally, use their developer API
    # Here we simulate some crime data per city
    # In real cases, you should register and use the FBI Crime Data API
    data = {
        "City": ["Los Angeles city, California", "San Diego city, California", "San Jose city, California"],
        "Crime_Rate_per_100k": [734.2, 542.1, 417.8]
    }
    return pd.DataFrame(data)

def merge_data():
    pop_df = fetch_population()
    income_df = fetch_income()
    crime_df = fetch_crime()
    merged = pd.merge(pop_df, income_df, on="City", how="left")
    merged = pd.merge(merged, crime_df, on="City", how="left")
    return merged

def save_to_excel(df, filename="us_data_report.xlsx"):
    df = df.sort_values(by="Population", ascending=False)
    df.reset_index(drop=True, inplace=True)

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="US Data", index=False)
        wb = writer.book
        ws = writer.sheets["US Data"]

        # Style des entêtes
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Ajouter un graphique
        plt.figure(figsize=(10, 6))
        df.head(10).plot(kind="bar", x="City", y="Population", title="Top 10 villes par population")
        plt.tight_layout()
        plt.savefig("population_chart.png")
        ws_img = wb.create_sheet("Charts")
        img = openpyxl.drawing.image.Image("population_chart.png")
        ws_img.add_image(img, "A1")

def main():
    df = merge_data()
    save_to_excel(df)
    print("Excel exporté avec population, revenus et taux de criminalité.")

if __name__ == "__main__":
    main()
